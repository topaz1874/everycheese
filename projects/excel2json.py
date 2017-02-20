# -*- coding: utf-8 -*-
import os
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "everycheese.config.settings")
 
import django
django.setup()

from django.core.serializers.json import DjangoJSONEncoder
from everycheese.warehouse.models import WarehouseItem
import json
import datetime
import openpyxl
import sys


def convert_data(filename):
	wb = openpyxl.load_workbook(filename)
	sheets = wb.get_sheet_names()
	# trans in sheet is 4 
	# trans out sheet is 5
	trans_in = sheets[4]
	trans_out = sheets[5]
	sheet = wb.get_sheet_by_name(sheets[4])
	# transList = []
	# for row in range(2, sheet.max_row+1):
	# 	transdata = {}
		# transdata['item'] = sheet['a'+ str(row)].value 
		# transdata['trans_date'] =sheet['b'+str(row)].value.strftime('%Y-%m-%d')
		# transdata['trans_amount']=sheet['d'+str(row)].value
		# transdata['trans_man'] =sheet['g'+str(row)].value
		
		# transList.append(transdata)
		
	# j = json.dumps(transList, cls=DjangoJSONEncoder)
	pk = WarehouseItem.objects.last.id
	names = [ _.name for _ in WarehouseItem.objects.all()]
	itemLst = []
	for row in range(2, sheet.max_row+1):
		itemDict = {}
		itemfields = {}		
		pk += 1
		itemfields['name'] = sheet['a' + str(row)].value
		itemfields['unit_price'] = sheet['e' + str(row)].value
		itemfields['amount'] = 0
		if itemfields['name'] not in names:
			itemDict['models'] = 'everycheese.warehouse' 
			itemDict['pk'] = pk
			itemDict['fields'] = itemfields

			itemLst.append(itemDict)

	j = json.dumps(itemLst, cls=DjangoJSONEncoder)	

	with open( '%s.json' % trans_in, 'w') as f:
		f.write(j)
			
if __name__ == '__main__':
    filename = sys.argv[1]
    convert_data(filename)